# -*- coding: utf-8 -*-
from __future__ import division
from __future__ import print_function
from __future__ import absolute_import
from __future__ import unicode_literals

import six
import shutil
import openpyxl
import datetime
from itertools import chain
from tempfile import TemporaryFile
from ..parser import Parser
from .. import exceptions
from .. import helpers


# Module API

class XLSXParser(Parser):
    """Parser to parse Excel modern `xlsx` data format.
    """

    # Public

    options = [
        'sheet',
        'fill_merged_cells',
        'preserve_formatting',
    ]

    def __init__(self, loader, force_parse=False, sheet=1,
            fill_merged_cells=False, preserve_formatting=False):
        self.__loader = loader
        self.__sheet_pointer = sheet
        self.__fill_merged_cells = fill_merged_cells
        self.__preserve_formatting = preserve_formatting
        self.__extended_rows = None
        self.__encoding = None
        self.__fragment = None
        self.__force_parse = force_parse
        self.__bytes = None

    @property
    def closed(self):
        return self.__bytes is None or self.__bytes.closed

    def open(self, source, encoding=None):
        self.close()
        self.__encoding = encoding
        self.__bytes = self.__loader.load(source, mode='b', encoding=encoding)

        # Create copy for remote source
        # For remote stream we need local copy (will be deleted on close by Python)
        # https://docs.python.org/3.5/library/tempfile.html#tempfile.TemporaryFile
        if getattr(self.__bytes, 'remote', False):
            new_bytes = TemporaryFile()
            shutil.copyfileobj(self.__bytes, new_bytes)
            self.__bytes.close()
            self.__bytes = new_bytes
            self.__bytes.seek(0)

        # Get book
        # To fill merged cells we can't use read-only because
        # `sheet.merged_cell_ranges` is not available in this mode
        self.__book = openpyxl.load_workbook(
            self.__bytes, read_only=not self.__fill_merged_cells, data_only=True)

        # Get sheet
        try:
            if isinstance(self.__sheet_pointer, six.string_types):
                self.__sheet = self.__book[self.__sheet_pointer]
            else:
                self.__sheet = self.__book.worksheets[self.__sheet_pointer - 1]
        except (KeyError, IndexError):
            message = 'Excel document "%s" doesn\'t have a sheet "%s"'
            raise exceptions.SourceError(message % (source, self.__sheet_pointer))
        self.__fragment = self.__sheet.title
        self.__process_merged_cells()

        # Reset parser
        self.reset()

    def close(self):
        if not self.closed:
            self.__bytes.close()

    def reset(self):
        helpers.reset_stream(self.__bytes)
        self.__extended_rows = self.__iter_extended_rows()

    @property
    def encoding(self):
        return self.__encoding

    @property
    def fragment(self):
        return self.__fragment

    @property
    def extended_rows(self):
        return self.__extended_rows

    # Private

    def __iter_extended_rows(self):
        for row_number, row in enumerate(self.__sheet.iter_rows(), start=1):
            yield (row_number, None, extract_row_values(row, self.__preserve_formatting))

    def __process_merged_cells(self):
        if self.__fill_merged_cells:
            for merged_cell_range in self.__sheet.merged_cells.ranges:
                merged_cell_range = str(merged_cell_range)
                self.__sheet.unmerge_cells(merged_cell_range)
                merged_rows = openpyxl.utils.rows_from_range(merged_cell_range)
                coordinates = list(chain(*merged_rows))
                value = self.__sheet[coordinates[0]].value
                for coordinate in coordinates:
                    cell = self.__sheet[coordinate]
                    cell.value = value


# Internal

NUMERIC_FORMATS = {
    '0': '{0:.0f}',
    '0.00': '{0:.2f}',
    '#,##0': '{0:,.0f}',
    '#,##0.00': '{0:,.2f}',
    '#,###.00': '{0:,.2f}',
}
TEMPORAL_FORMATS = {
    'm/d/yy': '%-m/%d/%y',
    'mm/dd/yy': '%m/%d/%y',
    'd-mmm': '%d-%b',
}


def extract_row_values(row, preserve_formatting=False):
    if preserve_formatting:
        values = []
        for cell in row:
            number_format = (cell.number_format or '').lower()
            number_format = number_format.replace('\\', '')
            numeric_format = NUMERIC_FORMATS.get(number_format)
            temporal_format = TEMPORAL_FORMATS.get(number_format)
            if isinstance(cell.value, (int, float)) and numeric_format:
                value = numeric_format.format(cell.value)
            elif isinstance(cell.value, datetime.datetime) and temporal_format:
                value = cell.value.strftime(temporal_format)
            else:
                value = cell.value
            values.append(value)
        return values
    return list(cell.value for cell in row)
